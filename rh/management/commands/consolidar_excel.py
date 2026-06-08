import os
import glob
import re
import openpyxl
import pandas as pd
from django.core.management.base import BaseCommand
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class Command(BaseCommand):
    help = 'Consolida competencias específicas respetando colores y textos sin omitir registros.'

    def limpiar_texto_clasificacion(self, texto):
        """Elimina números iniciales si existen (Ej: '4. TEXTO' -> 'TEXTO')"""
        if not texto:
            return ""
        texto_limpio = re.sub(r'^\s*\d+\s*[\.\s]*', '', str(texto))
        return texto_limpio.strip().upper()

    def handle(self, *args, **options):
        # Carpeta donde pones las evaluaciones de los empleados
        carpeta_origen = os.path.join(os.getcwd(), 'evaluaciones_excel')
        archivo_salida = 'Concentrado_Total_Competencias.xlsx'

        if not os.path.exists(carpeta_origen):
            self.stdout.write(self.style.ERROR(f"La carpeta '{carpeta_origen}' no existe."))
            return

        # Buscar archivos .xlsx o .xls
        archivos = glob.glob(os.path.join(carpeta_origen, "*.xlsx")) + glob.glob(os.path.join(carpeta_origen, "*.xls"))
        
        if not archivos:
            self.stdout.write(self.style.WARNING(f"No se encontraron archivos de Excel en '{carpeta_origen}'"))
            return

        datos_consolidados = []
        self.stdout.write("Iniciando extracción robusta basada en la estructura original...")

        for ruta_archivo in archivos:
            nombre_archivo = os.path.basename(ruta_archivo)
            # Limpiar nombre del empleado quitando la extensión
            empleado = nombre_archivo.replace(".xlsx", "").replace(".xls", "").strip().upper()
            
            try:
                # data_only=True para traer valores calculados y no las fórmulas de Excel
                wb_origen = openpyxl.load_workbook(ruta_archivo, data_only=True)
                # Seleccionamos la segunda pestaña (Competencias Específicas)
                ws_origen = wb_origen.worksheets[1]
            except Exception as e:
                self.stdout.write(self.style.ERROR(f"[-] Error al abrir {nombre_archivo}: {e}"))
                continue

            dentro_de_competencias = False
            clasificacion_actual = "GENERAL"

            for row in ws_origen.iter_rows(values_only=False):
                celda_a = row[0]
                val_a = str(celda_a.value).strip() if celda_a.value is not None else ""
                
                # Unimos el texto de toda la fila para evaluar palabras clave de parada o inicio
                texto_fila_completa = " ".join([str(c.value) for c in row if c.value is not None]).upper()

                # 1. Condición de Parada Absoluta
                if "FORTALEZAS" in texto_fila_completa and "OPORTUNIDAD" in texto_fila_completa:
                    break

                # 2. Condición de Inicio
                if "COMPETENCIAS A EVALUAR" in texto_fila_completa:
                    dentro_de_competencias = True
                    continue

                if not dentro_de_competencias or not val_a or val_a.lower() == "nan" or val_a == "":
                    continue

                # Omitir filas de escala, puntuaciones o comentarios del evaluador
                if val_a.startswith("*") or "PUNTOS" in texto_fila_completa or "PROMEDIO" in texto_fila_completa:
                    continue

                # 3. DETECCIÓN DE CLASIFICACIÓN (MÉTODO INFALIBLE POR COLOR)
                # Si la celda tiene color de fondo (Cualquier azul, gris, etc., diferente de vacío/blanco)
                fill = celda_a.fill
                es_clasificacion = False
                if fill and fill.fill_type and fill.fill_type != "none":
                    color_rgb = fill.start_color.rgb if fill.start_color else None
                    if color_rgb and color_rgb != "00000000" and color_rgb != "FFFFFFFF" and color_rgb != "FFFFFF":
                        es_clasificacion = True

                # Respaldo por si se rompe el estilo: si empieza con un patrón "Número. Espacio" (Ej: 4. FORTALECER...)
                if bool(re.match(r'^\s*\d+\s*\.', val_a)):
                    es_clasificacion = True

                if es_clasificacion:
                    # Guardamos la clasificación limpia para las competencias que vienen abajo
                    clasificacion_actual = self.limpiar_texto_clasificacion(val_a)
                    # Corregir posibles errores de dedo comunes del origen de datos
                    clasificacion_actual = clasificacion_actual.replace("INORMACION", "INFORMACIÓN")
                else:
                    # SI NO ES CLASIFICACIÓN, ES UNA COMPETENCIA/ACTIVIDAD VÁLIDA
                    # Filtramos que tenga una longitud mínima para evitar celdas con basura de un solo carácter
                    if len(val_a) > 5:
                        datos_consolidados.append({
                            "Clasificación": clasificacion_actual,
                            "Competencia": val_a,
                            "Empleado": empleado
                        })

        if not datos_consolidados:
            self.stdout.write(self.style.WARNING("No se logró extraer información. Revisa la estructura de los archivos."))
            return

        # 4. Construcción del Reporte Excel de Salida Organizado
        df_final = pd.DataFrame(datos_consolidados)
        wb_salida = openpyxl.Workbook()
        ws_salida = wb_salida.active
        ws_salida.title = "Matriz Unificada"
        ws_salida.views.sheetView[0].showGridLines = True

        # Estilos visuales del reporte maestro
        fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        font_title = Font(name="Arial", size=14, bold=True, color="1E293B")
        font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Arial", size=10, color="334155")
        border_thin = Border(left=Side(style="thin", color="E2E8F0"), right=Side(style="thin", color="E2E8F0"),
                             top=Side(style="thin", color="E2E8F0"), bottom=Side(style="thin", color="E2E8F0"))

        ws_salida["A1"] = "Concentrado Maestro de Competencias Específicas"
        ws_salida["A1"].font = font_title

        headers = ["Clasificación", "Competencia", "Empleado"]
        for col_num, title in enumerate(headers, 1):
            cell = ws_salida.cell(row=3, column=col_num, value=title)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center" if col_num == 3 else "left", vertical="center")

        row_num = 4
        for _, r_data in df_final.iterrows():
            cells = [
                ws_salida.cell(row=row_num, column=1, value=r_data["Clasificación"]),
                ws_salida.cell(row=row_num, column=2, value=r_data["Competencia"]),
                ws_salida.cell(row=row_num, column=3, value=r_data["Empleado"])
            ]
            for idx, c in enumerate(cells, 1):
                c.font = font_data
                c.border = border_thin
                c.alignment = Alignment(horizontal="center" if idx == 3 else "left", vertical="center", wrap_text=True)
                if row_num % 2 == 0:
                    c.fill = fill_zebra
            row_num += 1

        ws_salida.column_dimensions["A"].width = 45
        ws_salida.column_dimensions["B"].width = 85
        ws_salida.column_dimensions["C"].width = 30
        
        wb_salida.save(archivo_salida)
        self.stdout.write(self.style.SUCCESS(f"¡Hecho! Se procesaron {len(df_final)} registros. Archivo guardado como '{archivo_salida}'."))