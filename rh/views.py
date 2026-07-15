# ==========================================
# SECCIÓN DE IMPORTS CORREGIDA
# ==========================================
import openpyxl  
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  
from django.core.mail import send_mail
from django.db.models import Avg, Q 
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.apps import apps 
from django.http import HttpResponse
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth import login
from django.contrib import messages
from django.contrib import admin

# 💡 INCLUSIÓN: Importamos el nuevo modelo de la tabla intermedia
from .models import (
    Empleado, CompetenciaClasificacion, Competencia, Evaluacion, 
    EvaluacionDet, EvaluacionComentario, EmpleadoCompetenciaAsignada, TokenAccesoEvaluacion
)
from django.db import connection
from django.utils import timezone

@login_required
def panel_evaluacion_view(request, subordinado_id=None):
    try:
        usuario_logueado = Empleado.objects.get(user=request.user)
    except Empleado.DoesNotExist:
        messages.error(request, "Tu usuario no está vinculado a un registro de Empleado.")
        return redirect('admin:index')

    evaluacion_activa = Evaluacion.objects.filter().last()
    if not evaluacion_activa:
        context = {'error_mensaje': "No hay evaluaciones configuradas en este momento."}
        return render(request, 'evaluaciones/panel_evaluacion.html', context)

    if subordinado_id:
        empleado_a_evaluar = get_object_or_404(Empleado, id_empleado=subordinado_id)
        es_autoevaluacion = False
    else:
        empleado_a_evaluar = usuario_logueado
        es_autoevaluacion = True

    tipo_evaluador = 'E' if es_autoevaluacion else 'J'

    # =========================================================================
    # 1. EXTRACCIÓN INTELIGENTE DE COMPETENCIAS
    # =========================================================================
    competencias_globales_ids = list(Competencia.objects.filter(
        id_clasificacion__tipo='G'
    ).values_list('id_competencia', flat=True))

    competencias_asignadas_ids = list(EmpleadoCompetenciaAsignada.objects.filter(
        id_empleado=empleado_a_evaluar
    ).values_list('id_competencia_id', flat=True))

    ids_competencias_validas = list(set(competencias_globales_ids + competencias_asignadas_ids))

    competencias_reales = Competencia.objects.filter(
        id_competencia__in=ids_competencias_validas
    ).select_related('id_clasificacion')

    # =========================================================================
    # 2. RESPUESTAS PREVIAS Y CALIFICACIONES (LLAVES NUMÉRICAS ENTERAS)
    # =========================================================================
    respuestas_previos = EvaluacionDet.objects.filter(
        id_evaluacion=evaluacion_activa,
        id_empleado=empleado_a_evaluar,
        tipo=tipo_evaluador
    )
    ya_contestado = respuestas_previos.exists()
    
    # 💡 CORRECCIÓN DE NOMBRE: notas_guardadas (Con la "r" correspondiente)
    notas_guardadas = {}
    for r in respuestas_previos:
        if r.id_competencia_id is not None:
            # Guardamos el valor tal cual viene de la base de datos
            notas_guardadas[int(r.id_competencia_id)] = r.calificacion

    comentarios_previos = EvaluacionComentario.objects.filter(
        id_evaluacion=evaluacion_activa,
        id_empleado=empleado_a_evaluar,
        tipo_evaluador=tipo_evaluador
    )
    comentarios_dict = {c.tipo_bloque: c for c in comentarios_previos}
    comentario_g = comentarios_dict.get('G')
    comentario_e = comentarios_dict.get('E')

    # =========================================================================
    # SUBORDINADOS ASOCIADOS DIRECTAMENTE
    # =========================================================================
    subordinados_pendientes = []
    equipo = Empleado.objects.filter(id_jefe_id=usuario_logueado.id_empleado).exclude(id_empleado=usuario_logueado.id_empleado)
    
    for miembro in equipo:
        ya_evaluado_por_jefe = EvaluacionDet.objects.filter(
            id_evaluacion=evaluacion_activa,
            id_empleado=miembro,
            tipo='J'
        ).exists()
        
        subordinados_pendientes.append({
            'empleado': miembro,
            'estatus': 'Contestado' if ya_evaluado_por_jefe else 'Pendiente'
        })

    # =========================================================================
    # 3. CONSTRUCCIÓN DE LA ESTRUCTURA PARA EL HTML
    # =========================================================================
    mapa_clasificaciones = {}

    for comp in competencias_reales:
        clasif = comp.id_clasificacion
        if clasif.id_clasificacion not in mapa_clasificaciones:
            mapa_clasificaciones[clasif.id_clasificacion] = {
                'clasificacion': clasif,        
                'competencia_list': []          
            }
        mapa_clasificaciones[clasif.id_clasificacion]['competencia_list'].append(comp)

    clasificaciones_generales = []
    clasificaciones_especificas = []

    for clasif_id, item in mapa_clasificaciones.items():
        clasif_obj = item['clasificacion']
        lista_comps = item['competencia_list']
        
        # 💡 FILTRADO EN PYTHON: El promedio de la barra superior ignora los valores menores o iguales a 0 (NA)
        valores_calificaciones = [
            notas_guardadas[c.id_competencia] 
            for c in lista_comps if c.id_competencia in notas_guardadas and notas_guardadas[c.id_competencia] > 0
        ]
        promedio = sum(valores_calificaciones) / len(valores_calificaciones) if valores_calificaciones else 0
        item['promedio'] = round(promedio, 2)

        tipo_codigo = str(clasif_obj.tipo).strip().upper()

        if tipo_codigo == 'G':
            clasificaciones_generales.append(item)
        elif tipo_codigo == 'E':
            clasificaciones_especificas.append(item)

    clasificaciones_generales = sorted(clasificaciones_generales, key=lambda x: x['clasificacion'].descripcion)
    clasificaciones_especificas = sorted(clasificaciones_especificas, key=lambda x: x['clasificacion'].descripcion)

    context = {
        'evaluacion': evaluacion_activa,
        'evaluacion_cerrada': getattr(evaluacion_activa, 'cerrada', False), 
        'empleado': empleado_a_evaluar,         
        'usuario_logueado': usuario_logueado,   
        'clasificaciones_generales': clasificaciones_generales,
        'clasificaciones_especificas': clasificaciones_especificas,
        'ya_autoevaluado': ya_contestado, 
        'notas_guardadas': notas_guardadas, # 💡 Entregamos la variable con el nombre correcto
        'comentario_g': comentario_g,
        'comentario_e': comentario_e,
        'subordinados': subordinados_pendientes,
        'es_autoevaluacion': es_autoevaluacion,
    }
    return render(request, 'evaluaciones/panel_evaluacion.html', context)

@login_required
def guardar_evaluacion_view(request):
    if request.method == "POST":
        try:
            evaluador = Empleado.objects.get(user=request.user)
        except Empleado.DoesNotExist:
            messages.error(request, "Error: Tu usuario no está ligado a un empleado.")
            return redirect('admin:index')

        evaluado_id = request.POST.get("evaluado_id")
        evaluado = get_object_or_404(Empleado, id_empleado=evaluado_id)

        evaluacion_activa = Evaluacion.objects.filter().last()
        if not evaluacion_activa:
            messages.error(request, "No hay una evaluación activa en este momento.")
            return redirect('admin:index')

        if getattr(evaluacion_activa, 'cerrada', False):
            messages.error(request, "El periodo de evaluaciones ha sido cerrado. No se permiten más modificaciones.")
            return redirect('panel_evaluacion')

        if str(evaluador.id_empleado) == str(evaluado.id_empleado):
            tipo_evaluador = 'E'
        else:
            tipo_evaluador = 'J'

        # 1. GUARDAR RETROALIMENTACIÓN DE COMPETENCIAS GENERALES
        fortalezas_gen = request.POST.get("fortalezas_generales", "").strip()
        oportunidades_gen = request.POST.get("oportunidades_generales", "").strip()
        
        EvaluacionComentario.objects.update_or_create(
            id_evaluacion=evaluacion_activa,
            id_empleado=evaluado,
            tipo_bloque='G',       
            tipo_evaluador=tipo_evaluador,
            defaults={
                'fortalezas': Exam_clean_text(fortalezas_gen),
                'areas_oportunidad': Exam_clean_text(oportunidades_gen)
            }
        )

        # 2. GUARDAR RETROALIMENTACIÓN DE COMPETENCIAS ESPECÍFICAS
        fortalezas_esp = request.POST.get("fortalezas_especificas", "").strip()
        oportunidades_esp = request.POST.get("oportunidades_especificas", "").strip()
        
        EvaluacionComentario.objects.update_or_create(
            id_evaluacion=evaluacion_activa,
            id_empleado=evaluado,
            tipo_bloque='E',       
            tipo_evaluador=tipo_evaluador,
            defaults={
                'fortalezas': Exam_clean_text(fortalezas_esp),
                'areas_oportunidad': Exam_clean_text(oportunidades_esp)
            }
        )

        # 3. GUARDAR CALIFICACIONES NUMÉRICAS
        for key, value in request.POST.items():
            if key.startswith("nota_") and value:
                competencia_id = key.split("_")[1]
                competencia = get_object_or_404(Competencia, id_competencia=competencia_id)
                
                EvaluacionDet.objects.update_or_create(
                    id_evaluacion=evaluacion_activa,
                    id_empleado=evaluado,
                    id_competencia=competencia,
                    tipo=tipo_evaluador,
                    defaults={'calificacion': int(value)} 
                )

        nombre_evaluacion = evaluacion_activa.descripcion if evaluacion_activa.descripcion else "Evaluación de Desempeño"

        messages.success(
            request, 
            f'La evaluación "{nombre_evaluacion}" de {evaluado.nombre_largo} se guardó correctamente.'
        )
        
        return redirect('panel_evaluacion')

    return redirect('panel_evaluacion')
        
def Exam_clean_text(text):
    return text if text != "" else None

@login_required
def resumen_evaluaciones_view(request):
    try:
        usuario_logueado = Empleado.objects.get(user=request.user)
    except Empleado.DoesNotExist:
        messages.error(request, "Tu usuario no está vinculado a un registro de Empleado.")
        return redirect('admin:index')

    # Consulta directa con la columna 'evaluacion' integrada desde Supabase
    query = """
        SELECT 
            nombre_largo,
            auto_gen,
            auto_esp,
            eval_gen,
            eval_esp,
            evaluacion
        FROM vista_resumen_evaluaciones
        ORDER BY nombre_largo ASC;
    """

    tabla_datos = []
    nombre_evaluacion = "Evaluación"  # Valor por defecto

    with connection.cursor() as cursor:
        cursor.execute(query)
        rows = cursor.fetchall()

        for row in rows:
            nombre_emp = row[0]
            
            # Formateamos valores numéricos controlando los Nulos (None)
            auto_gen = float(row[1]) if row[1] is not None else 0.0
            auto_esp = float(row[2]) if row[2] is not None else 0.0
            eval_gen = float(row[3]) if row[3] is not None else 0.0
            eval_esp = float(row[4]) if row[4] is not None else 0.0
            
            # Recuperamos el nombre dinámico del periodo
            if row[5]:
                nombre_evaluacion = str(row[5])

            # --- PROMEDIO AUTOEVALUACIÓN ---
            if auto_gen > 0 and auto_esp > 0:
                promedio_auto = (auto_gen + auto_esp) / 2.0
            else:
                promedio_auto = auto_gen if auto_gen > 0 else auto_esp

            # --- PROMEDIO EVALUACIÓN DEL JEFE ---
            if eval_gen > 0 and eval_esp > 0:
                promedio_jefe = (eval_gen + eval_esp) / 2.0
            else:
                promedio_jefe = eval_gen if eval_gen > 0 else eval_esp

            # Omitir registros sin ninguna evaluación iniciada
            #if promedio_auto == 0.0 and promedio_jefe == 0.0:
            #    continue

            # --- PROMEDIO TOTAL COMBINADO ---
            if promedio_auto > 0 and promedio_jefe > 0:
                promedio_total = (promedio_auto + promedio_jefe) / 2.0
            else:
                promedio_total = promedio_auto if promedio_auto > 0 else promedio_jefe

            # --- NUEVA LÓGICA DE GRATIFICACIÓN (TEXTO EN RANGOS) ---
            if promedio_total is None or promedio_total == 0.0:
                gratificacion = "Sin evaluar"
            elif promedio_auto == 0 or promedio_jefe == 0:
                gratificacion = "Incompleta"
            elif 1.0 <= promedio_total <= 1.5:
                gratificacion = "0"
            elif 1.6 <= promedio_total <= 1.9:
                gratificacion = "15 días"
            elif 2.0 <= promedio_total <= 2.9:
                gratificacion = "1 mes"
            elif 3.0 <= promedio_total <= 3.9:
                gratificacion = "2 meses"
            elif 4.0 <= promedio_total <= 5.0:
                gratificacion = "3 meses"
            else:
                gratificacion = "Fuera de rango"

            tabla_datos.append({
                'empleado_nombre': nombre_emp,
                'promedio_auto': promedio_auto,
                'promedio_jefe': promedio_jefe,
                'promedio_total': promedio_total,
                'gratificacion': gratificacion
            })

    context = {
        'nombre_evaluacion': nombre_evaluacion,
        'tabla_datos': tabla_datos,
        'usuario_logueado': usuario_logueado,
    }
    return render(request, 'evaluaciones/resumen_evaluaciones.html', context)

@login_required
def descargar_plantilla_excel(request, model_name):
    try:
        model = apps.get_model('rh', model_name)
    except LookupError:
        return HttpResponse("Modelo no encontrado", status=404)

    # 1. Intentamos obtener la configuración de excel_columns de tu Admin personalizado
    columnas = []
    try:
        # Importamos localmente tu sitio de administración para evitar importaciones circulares
        from rh.admin import admin_site
        
        # Buscamos directamente en el registro de tu admin_site personalizado
        model_admin = admin_site._registry.get(model)
        
        if model_admin and hasattr(model_admin, 'excel_columns') and model_admin.excel_columns:
            # Quitamos los sufijos '_id' para que el Excel de la plantilla se descargue limpio 
            # con nombres como 'id_puesto' e 'id_departamento' y coincida con el importador inteligente
            columnas = [col.removesuffix('_id') for col in model_admin.excel_columns]
    except Exception as e:
        # Si algo falla, lo dejamos pasar para usar el respaldo de abajo
        pass

    # 2. Si no se encontró en tu admin_site o no tenía 'excel_columns', usamos el respaldo por defecto
    if not columnas:
        columnas = [field.name for field in model._meta.fields if field.name != 'id' and not field.auto_created]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Plantilla {model._meta.verbose_name_plural}"

    font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    fill_header = PatternFill(start_color='096446', end_color='096446', fill_type='solid')

    for col_num, column_title in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = font_header
        cell.fill = fill_header
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = max(len(column_title) + 5, 15)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=plantilla_{model_name}.xlsx'
    
    wb.save(response)
    return response

@login_required
def exportar_resumen_excel(request):
    # 1. Traer el texto de la barra de búsqueda si el usuario filtró en pantalla
    buscar_texto = request.GET.get('q', '').strip().lower()

    # 2. Ejecutar la CONSULTA REAL que sí existe en tu Supabase (vista_resumen_evaluaciones)
    query = """
        SELECT 
            nombre_largo,
            auto_gen,
            auto_esp,
            eval_gen,
            eval_esp,
            evaluacion
        FROM vista_resumen_evaluaciones
        ORDER BY nombre_largo ASC;
    """

    tabla_datos = []
    nombre_evaluacion = "Consolidado de Resultados"

    with connection.cursor() as cursor:
        cursor.execute(query)
        rows = cursor.fetchall()

        for row in rows:
            nombre_emp = row[0]
            
            # Si el usuario usó la barra de búsqueda en pantalla, filtramos aquí en memoria
            if buscar_texto and buscar_texto not in nombre_emp.lower():
                continue

            # Formateamos valores numéricos controlando los Nulos (None)
            auto_gen = float(row[1]) if row[1] is not None else 0.0
            auto_esp = float(row[2]) if row[2] is not None else 0.0
            eval_gen = float(row[3]) if row[3] is not None else 0.0
            eval_esp = float(row[4]) if row[4] is not None else 0.0
            
            if row[5]:
                nombre_evaluacion = str(row[5])

            # --- PROMEDIO AUTOEVALUACIÓN ---
            if auto_gen > 0 and auto_esp > 0:
                promedio_auto = (auto_gen + auto_esp) / 2.0
            else:
                promedio_auto = auto_gen if auto_gen > 0 else auto_esp

            # --- PROMEDIO EVALUACIÓN DEL JEFE ---
            if eval_gen > 0 and eval_esp > 0:
                promedio_jefe = (eval_gen + eval_esp) / 2.0
            else:
                promedio_jefe = eval_gen if eval_gen > 0 else eval_esp

            # Omitir registros sin ninguna evaluación iniciada (igual que en el HTML)
            #if promedio_auto == 0.0 and promedio_jefe == 0.0:
            #    continue

            # --- PROMEDIO TOTAL COMBINADO ---
            if promedio_auto > 0 and promedio_jefe > 0:
                promedio_total = (promedio_auto + promedio_jefe) / 2.0
            else:
                promedio_total = promedio_auto if promedio_auto > 0 else promedio_jefe

            # --- LÓGICA DE GRATIFICACIÓN (TEXTO EN RANGOS IDÉNTICO A TU PANTALLA) ---
            if promedio_total is None or promedio_total == 0.0:
                gratificacion = "Sin evaluar"
            elif promedio_auto == 0 or promedio_jefe == 0:
                gratificacion = "Incompleta"                    
            elif 1.0 <= promedio_total <= 1.5:
                gratificacion = "0"
            elif 1.6 <= promedio_total <= 1.9:
                gratificacion = "15 días"
            elif 2.0 <= promedio_total <= 2.9:
                gratificacion = "1 mes"
            elif 3.0 <= promedio_total <= 3.9:
                gratificacion = "2 meses"
            elif 4.0 <= promedio_total <= 5.0:
                gratificacion = "3 meses"
            else:
                gratificacion = "Fuera de rango"

            # Agregamos la fila procesada exactamente igual que en el HTML
            tabla_datos.append({
                'colaborador': nombre_emp,
                'promedio_auto': promedio_auto,
                'promedio_jefe': promedio_jefe,
                'promedio_total': promedio_total,
                'gratificacion': gratificacion
            })

    # ==========================================
    # 3. CONSTRUCCIÓN DEL ARCHIVO EXCEL (Openpyxl)
    # ==========================================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"
    ws.views.sheetView[0].showGridLines = True  # Asegura las líneas de cuadrícula

    # Estilos de diseño formal corporativo (Azul Marino del sistema)
    font_titulo = Font(name="Calibri", size=14, bold=True, color="1F497D")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Calibri", size=11, bold=False)
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_cebra = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    
    border_thin = Side(border_style="thin", color="D9D9D9")
    border_cell = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    # Título principal del Reporte
    ws.merge_cells("A1:E1")
    ws["A1"] = f"CONSOLIDADO: {nombre_evaluacion.upper()}"
    ws["A1"].font = font_titulo
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 35
    ws.append([]) # Fila 2 en blanco

    # Encabezados de Columnas que coinciden con tu tabla HTML
    headers = ["Colaborador", "Promedio Autoevaluación", "Promedio Evaluación Jefe", "Promedio Total", "Gratificación"]
    ws.append(headers)
    ws.row_dimensions[3].height = 26

    # Aplicar estilos a las cabeceras
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
        cell.border = border_cell

    # Llenar los datos renglón por renglón
    for idx, fila in enumerate(tabla_datos, start=4):
        ws.append([
            fila['colaborador'],
            round(fila['promedio_auto'], 2) if fila['promedio_auto'] > 0 else 0,
            round(fila['promedio_jefe'], 2) if fila['promedio_jefe'] > 0 else 0,
            round(fila['promedio_total'], 2),
            fila['gratificacion']
        ])
        
        ws.row_dimensions[idx].height = 20
        is_even = (idx % 2 == 0)

        # Estilizar las celdas de datos con efecto cebra
        for col_idx in range(1, 6):
            cell = ws.cell(row=idx, column=col_idx)
            cell.font = font_body
            cell.border = border_cell
            
            if is_even:
                cell.fill = fill_cebra
            
            # Alineación numérica o de texto
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

# =========================================================================
    # AUTOAJUSTAR EL ANCHO DE LAS COLUMNAS (CORREGIDO PARA CELDAS COMBINADAS)
    # =========================================================================
    from openpyxl.utils import get_column_letter

    for col in ws.columns:
        max_len = 0
        # Obtenemos de forma segura la letra de la columna actual
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            # Ignoramos la fila 1 por completo para evitar errores con las celdas combinadas del título
            if cell.row == 1:
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
                
        # Asignamos un ancho proporcional con un mínimo de 14
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    # 4. Enviar el archivo final al navegador
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="Consolidado_{nombre_evaluacion.replace(" ", "_")}.xlsx"'
    wb.save(response)
    
    return response

from openpyxl.utils import get_column_letter

@login_required
def exportar_detalle_competencias_excel(request):
    # 1. Capturar los filtros activos que vienen desde la pantalla HTML
    id_periodo = request.GET.get('periodo_id')
    buscar_texto = request.GET.get('q', '').strip().lower()

    # Si entran por primera vez, emular la lógica por defecto (traer la última evaluación)
    if not id_periodo:
        ultima_eval = Evaluacion.objects.filter().last()
        id_periodo = ultima_eval.id_evaluacion if ultima_eval else None

    # Intentar obtener la descripción de la evaluación para el encabezado del reporte
    nombre_evaluacion = "Consolidado de Detalles"
    if id_periodo:
        try:
            eval_obj = Evaluacion.objects.get(id_evaluacion=id_periodo)
            nombre_evaluacion = eval_obj.descripcion
        except Evaluacion.DoesNotExist:
            pass

    tabla_datos = []

    # 2. Consultar directamente tu nueva vista_evaluaciones
    if id_periodo:
        query = """
            SELECT 
                nombre_largo,
                departamento,
                puesto,
                clasificacion,
                competencia,
                evaluacion,
                autoevaluacion
            FROM vista_evaluaciones
            WHERE id_evaluacion = %s
            ORDER BY nombre_largo ASC, clasificacion ASC, competencia ASC;
        """

        with connection.cursor() as cursor:
            cursor.execute(query, [id_periodo])
            rows = cursor.fetchall()

            for row in rows:
                colaborador = row[0] or "Sin Nombre"
                departamento = row[1] or "Sin Área"
                puesto = row[2] or "Sin Puesto"
                clasificacion = row[3] or "Sin Clasificación"
                competencia = row[4] or "Sin Competencia"
                
                # Convertir calificaciones controlando nulos
                nota_emp = float(row[5]) if row[5] is not None else 0.0
                nota_jefe = float(row[6]) if row[6] is not None else 0.0

                # Aplicar el filtro de la barra de búsqueda en tiempo real si el usuario escribió algo
                if buscar_texto:
                    texto_fila = f"{colaborador} {departamento} {puesto} {clasificacion} {competencia}".lower()
                    if buscar_texto not in texto_fila:
                        continue

                tabla_datos.append({
                    'colaborador': colaborador,
                    'departamento': departamento,
                    'puesto': puesto,
                    'clasificacion': clasificacion,
                    'competencia': competencia,
                    'nota_emp': nota_emp,
                    'nota_jefe': nota_jefe
                })

    # ==========================================
    # 3. DISEÑO Y CONSTRUCCIÓN DEL EXCEL (Openpyxl)
    # ==========================================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalle de Calificaciones"
    ws.views.sheetView[0].showGridLines = True  # Mostrar cuadrícula explícita

    # Paleta de Estilos Corporativos (Azul institucional)
    font_titulo = Font(name="Calibri", size=14, bold=True, color="1F497D")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Calibri", size=11, bold=False)
    
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_cebra = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    fill_cero = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Resaltado amarillo claro para ausencias (0)
    
    border_thin = Side(border_style="thin", color="D9D9D9")
    border_cell = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    # Renglón 1: Título unificado
    ws.merge_cells("A1:G1")
    ws["A1"] = f"REPORTE DETALLADO POR COMPETENCIA: {nombre_evaluacion.upper()}"
    ws["A1"].font = font_titulo
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 35
    ws.append([]) # Renglón 2 en blanco

    # Renglón 3: Cabeceras de la tabla
    headers = ["Colaborador", "Departamento", "Puesto", "Clasificación", "Competencia", "Evaluación Empleado", "Evaluación Jefe"]
    ws.append(headers)
    ws.row_dimensions[3].height = 25

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center" if col_idx >= 6 else "left", vertical="center")
        cell.border = border_cell

    # Renglón 4 en adelante: Inyección de datos procesados
    for idx, fila in enumerate(tabla_datos, start=4):
        ws.append([
            fila['colaborador'],
            fila['departamento'],
            fila['puesto'],
            fila['clasificacion'],
            fila['competencia'],
            fila['nota_emp'],
            fila['nota_jefe']
        ])
        
        ws.row_dimensions[idx].height = 20
        is_even = (idx % 2 == 0)

        for col_idx in range(1, 8):
            cell = ws.cell(row=idx, column=col_idx)
            cell.font = font_body
            cell.border = border_cell
            
            # Aplicar cebra normal
            if is_even:
                cell.fill = fill_cebra
            
            # Alerta visual: si el empleado o el jefe tienen 0 (no han contestado), pintar en amarillo pastel
            if col_idx in [6, 7] and cell.value == 0:
                cell.fill = fill_cero

            # Formato y alineaciones por columna
            if col_idx >= 6:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = '0.0'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Autoajuste automático del ancho de columnas respetando celdas combinadas de la fila 1
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    # 4. Responder con el archivo binario directo al navegador
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="Detalle_Evaluaciones_{id_periodo}.xlsx"'
    wb.save(response)
    
    return response

@login_required
@user_passes_test(lambda u: u.is_staff, login_url='admin:index')
def asignacion_competencias_view(request):
    try:
        usuario_logueado = Empleado.objects.get(user=request.user)
    except Empleado.DoesNotExist:
        usuario_logueado = None

    # 1. Capturar los filtros enviados por el formulario (GET)
    search_query = request.GET.get('search', '').strip()
    filter_dept = request.GET.get('departamento', '').strip()
    filter_clasif = request.GET.get('clasificacion', '').strip()

    # 2. Consultas auxiliares para llenar los dropdowns de filtros dinámicamente
    listado_departamentos = []
    listado_clasificaciones = []
    
    with connection.cursor() as cursor:
        cursor.execute("SELECT DISTINCT departamento FROM vista_empleado_competencias WHERE departamento IS NOT NULL ORDER BY departamento;")
        listado_departamentos = [r[0] for r in cursor.fetchall()]
        
        cursor.execute("SELECT DISTINCT clasificacion FROM vista_empleado_competencias WHERE clasificacion IS NOT NULL ORDER BY clasificacion;")
        listado_clasificaciones = [r[0] for r in cursor.fetchall()]

    # 3. Construcción del Query Principal Dinámico con filtros WHERE
    base_query = """
        SELECT 
            id_empleado,
            nombre,
            departamento,
            clasificacion,
            competencia
        FROM vista_empleado_competencias
        WHERE 1=1
    """
    params = []

    # Aplicar filtro de búsqueda general (Nombre o Competencia)
    if search_query:
        base_query += " AND (nombre ILIKE %s OR competencia ILIKE %s)"
        params.extend([f"%{search_query}%", f"%{search_query}%"])

    # Aplicar filtro por Departamento
    if filter_dept:
        base_query += " AND departamento = %s"
        params.append(filter_dept)

    # Aplicar filtro por Clasificación
    if filter_clasif:
        base_query += " AND clasificacion = %s"
        params.append(filter_clasif)

    # Ordenamiento final idéntico
    base_query += " ORDER BY departamento ASC, nombre ASC;"

    competencias_asignadas = []

    with connection.cursor() as cursor:
        cursor.execute(base_query, params)
        rows = cursor.fetchall()

        for row in rows:
            competencias_asignadas.append({
                'id_empleado': row[0],
                'nombre_empleado': row[1],
                'departamento': row[2],
                'clasificacion': row[3],
                'competencia': row[4],
            })

    context = {
        'competencias_asignadas': competencias_asignadas,
        'usuario_logueado': usuario_logueado,
        # Enviar catálogos de filtros
        'listado_departamentos': listado_departamentos,
        'listado_clasificaciones': listado_clasificaciones,
        # Mantener los estados actuales seleccionados en la interfaz
        'search_query': search_query,
        'filter_dept': filter_dept,
        'filter_clasif': filter_clasif,
    }
    return render(request, 'evaluaciones/asignacion_competencias.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff, login_url='admin:index')
def exportar_competencias_excel(request):
    # 1. Capturar los mismos filtros aplicados en la pantalla
    search_query = request.GET.get('search', '').strip()
    filter_dept = request.GET.get('departamento', '').strip()
    filter_clasif = request.GET.get('clasificacion', '').strip()

    # 2. Construir la consulta SQL dinámica con los filtros
    base_query = """
        SELECT 
            id_empleado,
            nombre,
            departamento,
            clasificacion,
            competencia
        FROM vista_empleado_competencias
        WHERE 1=1
    """
    params = []

    if search_query:
        base_query += " AND (nombre ILIKE %s OR competencia ILIKE %s)"
        params.extend([f"%{search_query}%", f"%{search_query}%"])

    if filter_dept:
        base_query += " AND departamento = %s"
        params.append(filter_dept)

    if filter_clasif:
        base_query += " AND clasificacion = %s"
        params.append(filter_clasif)

    base_query += " ORDER BY departamento ASC, nombre ASC;"

    # 3. Crear el libro de Excel con Openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Competencias Asignadas"
    
    # Asegurar que se muestren las líneas de cuadrícula nativas de Excel
    ws.views.sheetView[0].showGridLines = True

    # 4. Estilos visuales institucionales (Verde Esmeralda Corporativo)
    font_titulo = Font(name="Arial", size=14, bold=True, color="096446")
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Arial", size=10, bold=False)
    font_mono = Font(name="Courier New", size=10, bold=False)
    
    fill_header = PatternFill(start_color="096446", end_color="096446", fill_type="solid")
    fill_cebra = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    border_delgado = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # Título del Reporte
    ws['A1'] = "Reporte de Competencias Asignadas por Colaborador"
    ws['A1'].font = font_titulo
    ws.row_dimensions[1].height = 25

    # Encabezados de la Tabla
    headers = ["ID Empleado", "Colaborador", "Departamento", "Clasificación", "Competencia / Descripción"]
    ws.append([]) # Fila 2 vacía para dar aire
    ws.append(headers) # Fila 3
    
    ws.row_dimensions[3].height = 24
    
    # Aplicar estilos a la cabecera
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center" if col_num == 1 else "left", vertical="center")
        cell.border = border_delgado

    # 5. Ejecutar consulta en la Base de Datos e inyectar registros
    with connection.cursor() as cursor:
        cursor.execute(base_query, params)
        rows = cursor.fetchall()

        row_index = 4
        for row in rows:
            ws.append([row[0], row[1], row[2], row[3], row[4]])
            ws.row_dimensions[row_index].height = 20
            
            # Aplicar efectos y cebra alternada a las celdas de datos
            for col_num in range(1, 6):
                cell = ws.cell(row=row_index, column=col_num)
                cell.font = font_mono if col_num == 1 else font_body
                cell.border = border_delgado
                
                # Alineaciones precisas
                if col_num == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Efecto cebra para las filas pares
                if row_index % 2 == 0:
                    cell.fill = fill_cebra
            
            row_index += 1

    # Autoajustar el ancho de las columnas según su contenido
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        # Omitir la primera fila del cálculo de longitud para que el título largo no rompa el ancho de la columna A
        for cell in col:
            if cell.row > 1 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Ancho fijo especial para la descripción de competencias para que no quede gigante
    ws.column_dimensions['E'].width = 60

    # 6. Preparar respuesta HTTP para descarga forzada de archivo Excel (.xlsx)
    response = HttpResponse(content_type='application/vnd.openpyxlsheet')
    response['Content-Disposition'] = 'attachment; filename="competencias_asignadas.xlsx"'
    
    wb.save(response)
    return response

# 🚨 QUITAMOS el @login_required de aquí arriba porque esta vista es libre para que el usuario se loguee
def acceso_magico_view(request, token_uuid):
    token = get_object_or_404(TokenAccesoEvaluacion, id_token=token_uuid)
    
    if token.es_valido():
        # Obtener el usuario de Django ligado a ese empleado
        usuario_django = token.empleado.user 
        
        if usuario_django:
            # 🌟 El truco: Iniciar sesión en Django sin pedir contraseña
            usuario_django.backend = 'django.contrib.auth.backends.ModelBackend'
            login(request, usuario_django)
            
            # Opcional: Marcar como utilizado si solo quieres que sirva una vez
            # token.utilizado = True
            # token.save()
            
            return redirect('panel_evaluacion') # 👈 Ruta de tu panel de evaluaciones
            
    # 🌟 Si el token expiró o falló, redirigimos al login del admin personalizado usando su namespace correcto
    messages.error(request, "El enlace de acceso ya no es válido o ha expirado.")
    return redirect('admin:login')