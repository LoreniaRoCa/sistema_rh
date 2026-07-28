from django.contrib.auth.models import User, Group
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.contrib.auth.admin import GroupAdmin as BaseGroupAdmin
from django.core.mail import send_mail
from django.db import models  
from django.contrib import admin
from django.urls import path
from django.shortcuts import render, redirect
from django.contrib import messages
from django import forms
import openpyxl
from unfold.admin import ModelAdmin, TabularInline  
from unfold.sites import UnfoldAdminSite  # ⬅️ ¡ESTA ES LA LÍNEA MÁGICA QUE FALTA!
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from django.db import connection
from .models import TokenAccesoEvaluacion, Empleado
import uuid
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from unfold.admin import ModelAdmin
from unfold.forms import ActionForm

class CustomActionForm(ActionForm):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if 'action' in self.fields:
            self.fields['action'].empty_label = "Seleccionar acción..."
            # Forzamos las opciones para borrar el "Select action" en inglés de la lista
            choices = [('', 'Seleccionar acción...')]
            for option_key, option_label in self.fields['action'].choices:
                if option_key != '':
                    choices.append((option_key, option_label))
            self.fields['action'].choices = choices

class CustomAdminSite(UnfoldAdminSite):
    index_template = "admin/index.html"

class CustomAdminSite(UnfoldAdminSite):
    index_template = "admin/index.html"

    def index(self, request, extra_context=None):
        extra_context = extra_context or {}
        
        departamentos_dict = {}
        jefes_dict = {} 
        
        total_global_empleados = 0
        total_ambas_completadas_global = 0
        total_auto_global = 0
        total_jefe_global = 0

        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    departamento, 
                    jefe, 
                    numempleados, 
                    autoevaluados, 
                    evaluados,
                    ambas  
                FROM rh_vista_dashboard_departamentos
            """)
            rows = cursor.fetchall()
            
            for row in rows:
                dep_nombre = row[0] or "Sin Departamento"
                jefe_nombre = row[1] or "Sin Jefe Asignado"
                
                num_emp = int(row[2] or 0)
                auto_ev = int(row[3] or 0)
                jefe_ev = int(row[4] or 0)
                ambas_ev = int(row[5] or 0)  

                # 1. MANTENER DEPARTAMENTOS INTACTOS
                if dep_nombre not in departamentos_dict:
                    departamentos_dict[dep_nombre] = {
                        'num_empleados': 0,
                        'auto_evaluados': 0,
                        'evaluados': 0,
                        'porcentaje': 0
                    }
                
                departamentos_dict[dep_nombre]['num_empleados'] += num_emp
                departamentos_dict[dep_nombre]['auto_evaluados'] += auto_ev
                departamentos_dict[dep_nombre]['evaluados'] += jefe_ev
                
                total_dep = departamentos_dict[dep_nombre]['num_empleados']
                autos_dep = departamentos_dict[dep_nombre]['auto_evaluados']
                jefes_dep = departamentos_dict[dep_nombre]['evaluados']
                
                # ➡️ CORRECCIÓN AQUÍ: El total de evaluaciones esperadas es el doble del número de empleados (Auto + Jefe)
                evaluaciones_totales_esperadas = total_dep * 2
                evaluaciones_realizadas = autos_dep + jefes_dep
                
                if evaluaciones_totales_esperadas > 0:
                    departamentos_dict[dep_nombre]['porcentaje'] = round((evaluaciones_realizadas / evaluaciones_totales_esperadas) * 100)
                else:
                    departamentos_dict[dep_nombre]['porcentaje'] = 0

                # 2. ACTUALIZACIÓN DE KPIS GLOBALES
                total_global_empleados += num_emp
                total_auto_global += auto_ev
                total_jefe_global += jefe_ev
                total_ambas_completadas_global += ambas_ev  
                
                # 3. AGRUPACIÓN Y SUMARIZACIÓN POR JEFE
                if jefe_nombre not in jefes_dict:
                    jefes_dict[jefe_nombre] = {
                        'nombre': jefe_nombre,
                        'total_empleados': 0,
                        'auto_contestadas': 0,
                        'jefe_contestadas': 0,
                    }
                
                jefes_dict[jefe_nombre]['total_empleados'] += num_emp
                jefes_dict[jefe_nombre]['auto_contestadas'] += auto_ev
                jefes_dict[jefe_nombre]['jefe_contestadas'] += jefe_ev

        # --- POST-PROCESAMIENTO DE JAFES ---
        jefes_data = []
        for jefe_nombre, data in jefes_dict.items():
            tot_emp = data['total_empleados']
            auto_cont = data['auto_contestadas']
            jefe_cont = data['jefe_contestadas']

            auto_pct = round((auto_cont / tot_emp * 100), 1) if tot_emp > 0 else 0.0
            jefe_pct = round((jefe_cont / tot_emp * 100), 1) if tot_emp > 0 else 0.0

            color_auto = "#ef4444" if auto_pct < 40 else ("#f59e0b" if auto_pct < 85 else "#72a651")
            color_jefe = "#ef4444" if jefe_pct < 40 else ("#f59e0b" if jefe_pct < 85 else "#72a651")

            jefes_data.append({
                'nombre': jefe_nombre,
                'total_empleados': tot_emp,
                'auto_contestadas': auto_cont,
                'auto_pct': auto_pct,
                'color_auto': color_auto,
                'jefe_contestadas': jefe_cont,
                'jefe_pct': jefe_pct,
                'color_jefe': color_jefe,
            })

        jefes_data = sorted(jefes_data, key=lambda x: x['nombre'])

        # --- CÁLCULO DE KPIS GLOBAL ---
        pct_general = round((total_ambas_completadas_global / total_global_empleados * 100), 1) if total_global_empleados > 0 else 0
        pct_auto = round((total_auto_global / total_global_empleados * 100), 1) if total_global_empleados > 0 else 0
        pct_jefe = round((total_jefe_global / total_global_empleados * 100), 1) if total_global_empleados > 0 else 0

        offset_general = round(251.2 * (1 - (pct_general / 100)), 1)
        offset_auto = round(251.2 * (1 - (pct_auto / 100)), 1)
        offset_jefe = round(251.2 * (1 - (pct_jefe / 100)), 1)

        extra_context["global_kpis"] = {
            "general_completadas": total_ambas_completadas_global,
            "general_porcentaje": pct_general,
            "general_offset": offset_general,
            "auto_contestadas": total_auto_global,
            "auto_porcentaje": pct_auto,
            "auto_offset": offset_auto,
            "jefe_contestadas": total_jefe_global,
            "jefe_porcentaje": pct_jefe,
            "jefe_offset": offset_jefe,
        }

        extra_context["departamentos"] = departamentos_dict
        extra_context["jefes_data"] = jefes_data

        return super().index(request, extra_context=extra_context)

# Instanciación del sitio personalizado
admin_site = CustomAdminSite(name='custom_admin')

# Desasociamos los registros por defecto (en caso de ser necesario)
# e incorporamos el soporte de visualización para tu panel personalizado de Unfold
try:
    admin.site.unregister(User)
    admin.site.unregister(Group)
except admin.sites.NotRegistered:
    pass

# Registramos Usuarios y Grupos para que tu URLconf sepa que existen en /admin/
# Registramos Usuarios y Grupos para que tu URLconf sepa que existen en tu 'admin_site' personalizado de Unfold
@admin.register(User, site=admin_site)  # 🌟 Agregamos 'site=admin_site' para que se vinculen a tu dashboard
class UserAdmin(BaseUserAdmin, ModelAdmin):
    pass

@admin.register(Group, site=admin_site) # 🌟 Agregamos 'site=admin_site' para que se vinculen a tu dashboard
class GroupAdmin(BaseGroupAdmin, ModelAdmin):
    pass


# Importación explícita de todos los modelos requeridos
from .models import (
    Puesto, Departamento, Empleado, CompetenciaClasificacion, 
    Competencia, Evaluacion, EvaluacionDet, ClasificacionPorPuesto, 
    ClasificacionPorEmpleado, EmpleadoCompetenciaAsignada
)



class CatalogosOrdenadosAdmin(admin.ModelAdmin):
    """
    Clase base para que todos los combos (ForeignKeys) del sistema
    se ordenen alfabéticamente de forma automática, sin alterar el listado principal.
    """
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        # 1. Obtenemos el modelo al que apunta esta clave heredada
        target_model = db_field.remote_field.model
        campos = [f.name for f in target_model._meta.fields]

        # 2. Si el modelo tiene un campo 'nombre_largo' o 'descripcion', ordenamos por él
        if 'nombre_largo' in campos:
            kwargs["queryset"] = target_model.objects.order_by('nombre_largo')
        elif 'descripcion' in campos:
            kwargs["queryset"] = target_model.objects.order_by('descripcion')
        elif 'nombre' in campos:
            kwargs["queryset"] = target_model.objects.order_by('nombre')

        return super().formfield_for_foreignkey(db_field, request, **kwargs)
class ExcelUploadForm(forms.Form):
    archivo_excel = forms.FileField(label="Selecciona el archivo de Excel (.xlsx)")


# =========================================================================
#  CLASE BASE PARA IMPORTACIÓN EXCEL
# =========================================================================
# =========================================================================
#  CLASE BASE PARA IMPORTACIÓN EXCEL (CORREGIDA)
# =========================================================================
# =========================================================================
#  CLASE BASE PARA IMPORTACIÓN EXCEL (MÉTODO EXPORTAR INTEGRADO)
# =========================================================================
class ExcelImportAdmin(ModelAdmin):
    import_template = "admin/importar_excel.html"
    change_list_template = "admin/carga_masiva_change_list.html"
    # 🌟 APLICADO GLOBALMENTE: Todos los catálogos heredarán esto automáticamente
    action_form = CustomActionForm
    action_submit_label = "Ejecutar"
   
    model_class = None       
    pk_field_name = None     
    excel_columns = []       

    list_per_page = 25
    list_select_related = True

    # 1. MANTENIDO: Generador de URLs para los botones del template personalizado
    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        
        if self.model:
            app_label = self.model._meta.app_label
            model_name = self.model._meta.model_name
            query_string = request.GET.urlencode()
            
            # Apunta al get_urls interno del Admin
            url_exportar = f"/admin/{app_label}/{model_name}/exportar-excel/"
            if query_string:
                url_exportar += f"?{query_string}"
                
            extra_context['url_exportar_excel'] = url_exportar
            
        return super().changelist_view(request, extra_context=extra_context)

    # 2. MANTENIDO: Tus botones de acción rápidos (Editar/Eliminar)
    def acciones_rh(self, obj):
        app_label = obj._meta.app_label
        model_name = obj._meta.model_name
        editar_url = f"/admin/{app_label}/{model_name}/{obj.pk}/change/"
        eliminar_url = f"/admin/{app_label}/{model_name}/{obj.pk}/delete/"
        
        return format_html(
            '<a href="{}" title="Editar" style="'
            'display: inline-flex !important; align-items: center !important; justify-content: center !important; '
            'width: 26px !important; height: 26px !important; border-radius: 4px !important; margin-right: 6px !important; '
            'background-color: #72a651 !important; color: #ffffff !important; font-weight: bold !important; '
            'text-decoration: none !important; font-size: 14px !important; opacity: 1 !important; line-height: 1 !important;">'
            '&#9998;'
            '</a>'
            '<a href="{}" title="Eliminar" style="'
            'display: inline-flex !important; align-items: center !important; justify-content: center !important; '
            'width: 26px !important; height: 26px !important; border-radius: 4px !important; '
            'background-color: #de3a3a !important; color: #ffffff !important; font-weight: bold !important; '
            'text-decoration: none !important; font-size: 13px !important; opacity: 1 !important; line-height: 1 !important;">'
            '&#10006;'
            '</a>',
            editar_url, eliminar_url
        )
    acciones_rh.short_description = "Acciones"

    # 3. ACTUALIZADO: Registra tanto tu importador original como el nuevo exportador
    def get_urls(self):
        urls = super().get_urls()
        if not self.model:
            return urls

        app_label = self.model._meta.app_label
        model_name = self.model._meta.model_name

        custom_urls = [
            # Tu ruta de importación original
            path(
                'importar-excel/', 
                self.admin_site.admin_view(self.import_excel_view), 
                name=f'{app_label}_{model_name}_import_excel'
            ),
            # La nueva ruta segura de exportación integrada
            path(
                'exportar-excel/', 
                self.admin_site.admin_view(self.exportar_catalogo_view), 
                name=f'{app_label}_{model_name}_exportar_excel'
            ),
        ]
        return custom_urls + urls

    # 4. MANTENIDO AL 100%: Tu motor original de procesamiento de carga masiva
    def import_excel_view(self, request):
        if request.method == "POST":
            excel_file = request.FILES.get("excel_file")
            
            if not excel_file:
                messages.error(request, "Por favor, selecciona un archivo válido.")
                return redirect(f"/admin/{self.model_class._meta.app_label}/{self.model_class._meta.model_name}/")

            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, "El archivo debe ser un Excel (.xlsx o .xls).")
                return redirect(f"/admin/{self.model_class._meta.app_label}/{self.model_class._meta.model_name}/")

            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                sheet = wb.active

                header_row = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in sheet[1]]
                col_map = {name: idx for idx, name in enumerate(header_row) if name}

                success_count = 0
                error_count = 0

                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row):  
                        continue

                    data = {}
                    for col_name in self.excel_columns:
                        posibles_nombres = [
                            col_name.lower(),
                            col_name.lower().removesuffix('_id'),
                            col_name.lower() + '_id'
                        ]
                        
                        idx = None
                        for nombre in posibles_nombres:
                            if nombre in col_map:
                                idx = col_map[nombre]
                                break

                        if idx is not None and idx < len(row):
                            val = row[idx]
                            if isinstance(val, str):
                                val = val.strip()
                            data[col_name] = val

                    pk_lower = self.pk_field_name.lower()
                    pk_idx = col_map.get(pk_lower) or col_map.get(pk_lower.removesuffix('_id')) or col_map.get(pk_lower + '_id')
                    pk_value = row[pk_idx] if pk_idx is not None and pk_idx < len(row) else None

                    if pk_value is not None and str(pk_value).strip() != "":
                        try:
                            pk_value = int(float(str(pk_value).strip()))
                        except (ValueError, TypeError):
                            error_count += 1
                            messages.warning(request, f"Error en fila {row_idx}: El ID debe ser un número entero. Se obtuvo: {pk_value}")
                            continue
                    else:
                        pk_value = None

                    if pk_value:
                        try:
                            defaults_data = {k: v for k, v in data.items() if k != self.pk_field_name}
                            instance, created = self.model_class.objects.update_or_create(
                                **{self.pk_field_name: pk_value},
                                defaults=defaults_data
                            )
                            success_count += 1
                        except Exception as e:
                            error_count += 1
                            messages.warning(request, f"Error en fila {row_idx} al procesar ID {pk_value}: {e}")
                    else:
                        try:
                            self.model_class.objects.create(**data)
                            success_count += 1
                        except Exception as e:
                            error_count += 1
                            messages.warning(request, f"Error en fila {row_idx} (Sin ID): {e}")

                messages.success(request, f"Importación completada. Registros procesados: {success_count}. Errores: {error_count}")
                
            except Exception as e:
                messages.error(request, f"Error crítico al procesar el archivo: {e}")
                
        return redirect(f"/admin/{self.model_class._meta.app_label}/{self.model_class._meta.model_name}/")

    # 5. NUEVO: Agregado al final sin pisar nada. Exporta la data en base a tus columnas mapeadas
    def exportar_catalogo_view(self, request):
        if not self.model:
            raise Http404("Modelo no configurado.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Catálogo {self.model._meta.model_name.capitalize()}"

        # Usa las columnas que tú definiste para armar la cabecera exacta de tu plantilla
        columnas = self.excel_columns if self.excel_columns else [field.name for field in self.model._meta.fields]
        ws.append(columnas)

        # Rellenamos el reporte
        queryset = self.model.objects.all()
        for objeto in queryset:
            fila = [getattr(objeto, col, "") for col in columnas]
            fila_limpia = [str(val) if val is not None else "" for val in fila]
            ws.append(fila_limpia)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="catalogo_{self.model._meta.model_name}.xlsx"'
        wb.save(response)
        return response


# ==========================================
#   INLINE: CLASIFICACIÓN POR PUESTO
# ==========================================
class ClasificacionPorPuestoInlineForm(forms.ModelForm):
    class Meta:
        model = ClasificacionPorPuesto
        fields = ('id_clasificacion',)
        widgets = {
            'id_clasificacion': forms.Select(attrs={
                'style': 'width: 100% !important; min-width: 100% !important; height: 38px !important; display: block !important;'
            })
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        
        # 🌟 FILTRO MÁGICO: Limitamos las opciones del combo a solo las específicas ('E') ordenadas alfabéticamente
        self.fields['id_clasificacion'].queryset = CompetenciaClasificacion.objects.filter(tipo='E').order_by('descripcion')

        if self.instance and self.instance.pk:
            valor_actual = self.instance.id_clasificacion_id
            self.fields['id_clasificacion'].widget.value = valor_actual
            self.initial['id_clasificacion'] = valor_actual
            self.instance.__class__.__str__ = lambda self: ""
            self.instance.id_clasificacion = None
        else:
            self.initial['id_clasificacion'] = ""


class ClasificacionPorPuestoInline(TabularInline):
    model = ClasificacionPorPuesto
    form = ClasificacionPorPuestoInlineForm
    extra = 1  # 💡 Cambiado a 1 para que se vea más limpio en el panel de Unfold
    fields = ('id_clasificacion',)


# ==========================================
#   INLINE: CLASIFICACIÓN POR EMPLEADO
# ==========================================
class ClasificacionPorEmpleadoInlineForm(forms.ModelForm):
    nueva_competencia_texto = forms.CharField(
        required=False,
        label="Competencia Exclusiva (Captura libre)",
        widget=forms.TextInput(attrs={
            'style': 'width: 100% !important; min-width: 250px !important; height: 38px !important; padding: 0 10px !important;',
            'placeholder': 'Escribe aquí la competencia solo para este empleado...'
        })
    )

    class Meta:
        model = ClasificacionPorEmpleado
        fields = ('id_clasificacion', 'nueva_competencia_texto', 'motivo')
        widgets = {
            'id_clasificacion': forms.Select(attrs={
                'style': 'width: 100% !important; min-width: 200px !important; height: 38px !important;'
            }),
            'motivo': forms.TextInput(attrs={
                'style': 'width: 100% !important; min-width: 150px !important; height: 38px !important;'
            })
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.instance and self.instance.pk and self.instance.competencia_exclusiva_id:
            self.initial['nueva_competencia_texto'] = self.instance.competencia_exclusiva.descripcion
        
        if self.instance and self.instance.pk:
            valor_actual = self.instance.id_clasificacion_id
            self.fields['id_clasificacion'].widget.value = valor_actual
            self.initial['id_clasificacion'] = valor_actual
            self.instance.__class__.__str__ = lambda self: ""
            self.instance.id_clasificacion = None
        else:
            self.initial['id_clasificacion'] = ""

    def save(self, commit=True):
        instance = super().save(commit=False)
        texto_competencia = self.cleaned_data.get('nueva_competencia_texto')
        id_clasif = self.cleaned_data.get('id_clasificacion') or instance.id_clasificacion

        if texto_competencia and id_clasif:
            from .models import Competencia
            if instance.competencia_exclusiva:
                comp = instance.competencia_exclusiva
                comp.descripcion = texto_competencia
                comp.id_clasificacion = id_clasif
                comp.save()
            else:
                comp = Competencia.objects.create(
                    id_clasificacion=id_clasif,
                    descripcion=texto_competencia
                )
                instance.competencia_exclusiva = comp
        elif not texto_competencia:
            instance.competencia_exclusiva = None

        if commit:
            instance.save()
        return instance


class ClasificacionPorEmpleadoInline(TabularInline):
    model = ClasificacionPorEmpleado
    form = ClasificacionPorEmpleadoInlineForm
    extra = 1
    fields = ('id_clasificacion', 'nueva_competencia_texto', 'motivo')


# ==========================================
#   INLINE: COMPETENCIAS
# ==========================================
class CompetenciaInlineForm(forms.ModelForm):
    descripcion = forms.CharField(
        widget=forms.TextInput(attrs={
            'style': 'width: 100% !important; min-width: 100% !important; height: 38px !important; display: block !important; padding: 0 10px !important;'
        }),
        required=False,
    )

    class Meta:
        model = Competencia
        fields = ('descripcion',)

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.instance and self.instance.pk:
            valor_actual = self.instance.descripcion
            self.fields['descripcion'].widget.value = valor_actual
            self.initial['descripcion'] = valor_actual
            self.instance.descripcion = ""  
        else:
            self.initial['descripcion'] = ""


class CompetenciaInline(TabularInline):
    model = Competencia
    form = CompetenciaInlineForm
    extra = 3  
    fields = ('descripcion',)


# =========================================================================
#   FORMULARIO PRINCIPAL: MATRIZ CHECKLIST DE SELECCIÓN INTELIGENTE
# =========================================================================
class EmpleadoAdminForm(forms.ModelForm):
    competencias_seleccionadas = forms.ModelMultipleChoiceField(
        queryset=Competencia.objects.none(),
        widget=forms.MultipleHiddenInput(),
        required=False,
        label=""
    )

    class Meta:
        model = Empleado
        fields = '__all__'

        # 🌟 AGREGA ESTA PROPIEDAD AQUÍ ABAJO:
        widgets = {
            'fechaalta': forms.DateInput(
                format='%Y-%m-%d',  # 🌟 Obliga a Django a enviar el dato en formato AAAA-MM-DD
                attrs={'type': 'date'}
            ),
        }    

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        # 🌟 NUEVO: FORZAR ORDENAMIENTO ALFABÉTICO EN LOS COMBOS PRINCIPALES DEL FORMULARIO
        if 'user' in self.fields:
            self.fields['user'].queryset = self.fields['user'].queryset.order_by('username')
        
        if 'id_jefe' in self.fields:
            self.fields['id_jefe'].queryset = Empleado.objects.order_by('nombre_largo')
            
        if 'id_puesto' in self.fields:
            self.fields['id_puesto'].queryset = Puesto.objects.order_by('descripcion')
            
        if 'id_departamento' in self.fields:
            self.fields['id_departamento'].queryset = Departamento.objects.order_by('descripcion')

        if self.instance and self.instance.pk:
            empleado_id = self.instance.pk
            # 1. Obtenemos el ID del puesto de la base de datos física de forma segura
            puesto_id = self.instance.id_puesto_id

            # --- UNIVERSO A: Competencias que le corresponden por su Puesto ---
            clasif_puesto_ids = list(ClasificacionPorPuesto.objects.filter(
                id_puesto_id=puesto_id
            ).values_list('id_clasificacion_id', flat=True))

            competencias_del_puesto_ids = list(Competencia.objects.filter(
                id_clasificacion_id__in=clasif_puesto_ids
            ).values_list('id_competencia', flat=True))

            # --- UNIVERSO B: Competencias que el empleado ya tiene guardadas ---
            competencias_guardadas_ids = list(EmpleadoCompetenciaAsignada.objects.filter(
                id_empleado_id=empleado_id
            ).values_list('id_competencia_id', flat=True))

            # --- UNIFICACIÓN DE AMBOS UNIVERSOS (Igual al UNION de tu SQL) ---
            todas_competencias_ids = list(set(competencias_del_puesto_ids + competencias_guardadas_ids))

            # 2. Reconstruimos el queryset final limpio y ordenado
            queryset_competencias = Competencia.objects.filter(
                id_competencia__in=todas_competencias_ids
            ).select_related('id_clasificacion').order_by('id_clasificacion__descripcion', 'descripcion')

            # 3. Asignamos de forma explícita al campo para desactivar el candado de Django
            self.fields['competencias_seleccionadas'].queryset = queryset_competencias
            self.fields['competencias_seleccionadas'].initial = competencias_guardadas_ids

            # Guardamos para change_view
            self.queryset_competencias_custom = queryset_competencias
            self.competencias_iniciales_custom = competencias_guardadas_ids
        else:
            self.queryset_competencias_custom = Competencia.objects.none()
            self.competencias_iniciales_custom = []


# =========================================================================
#   REGISTRO DE CATÁLOGOS EN EL ADMINISTRADOR
# =========================================================================

class PuestoAdmin(ExcelImportAdmin):
    model_class = Puesto
    pk_field_name = 'id_puesto'
    excel_columns = ['id_puesto', 'descripcion']
    list_display = ('id_puesto', 'descripcion', 'acciones_rh')
    search_fields = ('descripcion',)
    inlines = [ClasificacionPorPuestoInline]

#admin.site.register(Puesto, PuestoAdmin)
admin_site.register(Puesto, PuestoAdmin)

class DepartamentoAdmin(ExcelImportAdmin):
    model_class = Departamento
    pk_field_name = 'id_departamento'
    excel_columns = ['id_departamento', 'descripcion']
    list_display = ('id_departamento', 'descripcion', 'acciones_rh')
    search_fields = ('descripcion',)

# @admin.action(description='Enviar Enlaces de evaluación por Correo')
# def enviar_enlaces_magicos(modeladmin, request, queryset):
#     # 1. De los empleados QUE SELECCIONASTE en la lista, filtramos los que tienen correo válido
#     empleados_seleccionados = queryset.exclude(CorreoElectronico__isnull=True).exclude(CorreoElectronico='')
    
#     # 2. Obtenemos los IDs de todos los que son jefes en todo el sistema
#     jefes_ids = Empleado.objects.exclude(id_jefe__isnull=True).values_list('id_jefe', flat=True).distinct()
    
#     # 3. Aplicamos el filtro OR pero únicamente sobre los empleados SELECCIONADOS en el Admin
#     empleados_a_procesar = empleados_seleccionados.filter(
#         Q(se_evalua=True) | Q(id_empleado__in=jefes_ids)
#     )
    
#     if not empleados_a_procesar.exists():
#         modeladmin.message_user(
#             request, 
#             "Ninguno de los empleados seleccionados cumple con las condiciones (se_evalua=True o ser Jefe) o no tienen correo válido.", 
#             messages.WARNING
#         )
#         return

#     # 4. Guardamos los IDs de los seleccionados válidos en la sesión
#     empleados_ids = list(empleados_a_procesar.values_list('id_empleado', flat=True).distinct())
    
#     session_key = f"envio_correos_{uuid.uuid4().hex}"
#     request.session[session_key] = {
#         'empleados_ids': empleados_ids,
#         'procesados': 0,
#         'total': len(empleados_ids),
#         'app_label': queryset.model._meta.app_label,
#         'model_name': queryset.model._meta.model_name,
#     }

#     return redirect(f"/admin/procesar-evaluaciones-loading/{session_key}/")

@admin.action(description='Enviar Enlaces de evaluación por Correo (A todo el personal aplicable)')
def enviar_enlaces_magicos(modeladmin, request, queryset):
    # 1. Tomamos la evaluación seleccionada en la lista
    evaluacion_seleccionada = queryset.first()
    if not evaluacion_seleccionada:
        modeladmin.message_user(request, "Por favor, selecciona una evaluación.", messages.ERROR)
        return

    # 2. Obtenemos todos los empleados del sistema con correo electrónico válido
    empleados_con_correo = Empleado.objects.exclude(CorreoElectronico__isnull=True).exclude(CorreoElectronico='')

    # 3. Identificamos quiénes son jefes (los que aparecen en el campo 'id_jefe' de otros empleados)
    jefes_ids = Empleado.objects.exclude(id_jefe__isnull=True).values_list('id_jefe', flat=True).distinct()

    # 4. Filtramos: que se evalúen O que sean jefes de alguien
    empleados_a_procesar = empleados_con_correo.filter(
        Q(se_evalua=True) | Q(id_empleado__in=jefes_ids)
    )

    if not empleados_a_procesar.exists():
        modeladmin.message_user(
            request, 
            "Ningún empleado cumple con las condiciones (se_evalua=True o ser Jefe) o no tienen correo válido.", 
            messages.WARNING
        )
        return

    # 5. Guardamos la lista de IDs a procesar en la sesión para la pantalla de carga (Loading)
    empleados_ids = list(empleados_a_procesar.values_list('id_empleado', flat=True).distinct())
    
    session_key = f"envio_correos_{uuid.uuid4().hex}"
    request.session[session_key] = {
        'empleados_ids': empleados_ids,
        'evaluacion_id': evaluacion_seleccionada.id_evaluacion, # Asociamos la evaluación seleccionada
        'procesados': 0,
        'total': len(empleados_ids),
        'app_label': 'rh',
        'model_name': 'evaluacion',
    }

    return redirect(f"/admin/procesar-evaluaciones-loading/{session_key}/")


# =========================================================================
# ACCIÓN 2: Para el Modelo "Empleado" (Catálogo de Empleados)
# =========================================================================
@admin.action(description='Enviar Enlaces de evaluación a los empleados seleccionados')
def enviar_enlaces_seleccionados(modeladmin, request, queryset):
    # 1. Filtramos del queryset seleccionado los que tienen correo electrónico válido
    empleados_a_procesar = queryset.exclude(CorreoElectronico__isnull=True).exclude(CorreoElectronico='')

    if not empleados_a_procesar.exists():
        modeladmin.message_user(
            request, 
            "Ninguno de los empleados seleccionados tiene un correo electrónico válido registrado.", 
            messages.WARNING
        )
        return

    # 2. Obtenemos la última evaluación configurada en el sistema para asociar los enlaces
    evaluacion_activa = Evaluacion.objects.filter().last()
    if not evaluacion_activa:
        modeladmin.message_user(
            request, 
            "No se puede realizar el envío porque no hay ninguna evaluación activa configurada en el sistema.", 
            messages.ERROR
        )
        return

    # 3. Guardamos los IDs de los empleados seleccionados en la sesión
    empleados_ids = list(empleados_a_procesar.values_list('id_empleado', flat=True).distinct())
    
    session_key = f"envio_correos_{uuid.uuid4().hex}"
    request.session[session_key] = {
        'empleados_ids': empleados_ids,
        'evaluacion_id': evaluacion_activa.id_evaluacion, # Asociamos la evaluación activa por defecto
        'procesados': 0,
        'total': len(empleados_ids),
        'app_label': 'rh',
        'model_name': 'empleado',
    }

    return redirect(f"/admin/procesar-evaluaciones-loading/{session_key}/")

# 🌟 VISTA INTERMEDIA: PANTALLA DE CARGA DE FRUVER PROCESADA POR AJAX EN TIEMPO REAL
def procesar_evaluaciones_loading_view(request, session_key):
    data = request.session.get(session_key)
    if not data:
        return redirect('/admin/')

    # Si es una petición AJAX (Fetch), procesamos un lote de 3 correos para no saturar
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'ajax' in request.GET:
        ids_restantes = data['empleados_ids']
        total_a_procesar = min(3, len(ids_restantes)) # Lotes de 3 en 3
        lote = ids_restantes[:total_a_procesar]
        
        contador_local = 0
        dominio_sitio = request.build_absolute_uri('/')
        
        for emp_id in lote:
            try:
                empleado = Empleado.objects.get(id_empleado=emp_id)
                token = TokenAccesoEvaluacion.objects.create(empleado=empleado)
                url_acceso = f"{dominio_sitio}evaluacion/acceso/{token.id_token}/"
                
                asunto = "Acceso Exclusivo: Tu Evaluación de Desempeño"
                mensaje = f"Hola {empleado.nombre_largo if hasattr(empleado, 'nombre_largo') else empleado.nombre_largo},\n\n" \
                          f"Te compartimos tu enlace personalizado para ingresar al sistema de evaluaciones de desempeño.\n" \
                          f"A través de este enlace podrás realizar tu autoevaluación (si te corresponde) y/o evaluar a tu personal a cargo.\n\n" \
                          f"Haz clic en el siguiente enlace para ingresar directamente sin necesidad de contraseña:\n" \
                          f"{url_acceso}\n\n" \
                          f"Este enlace expirará en 5 días.\n\n" \
                          f"Saludos cordiales,\nRecursos Humanos."
                
                send_mail(
                    asunto, mensaje, 'l.rodriguez@fruver.com.mx',
                    [empleado.CorreoElectronico], fail_silently=False
                )
                contador_local += 1
            except Exception:
                pass
        
        # Actualizamos el estado de la sesión
        data['empleados_ids'] = ids_restantes[total_a_procesar:]
        data['procesados'] += total_a_procesar
        request.session[session_key] = data
        
        # Si terminamos todo, mandamos el mensaje final de éxito
        if len(data['empleados_ids']) == 0:
            messages.success(request, f"Se generaron los tokens y se enviaron {data['procesados']} correos exitosamente.")
            del request.session[session_key] # Limpiamos sesión
            
        return JsonResponse({
            'completado': len(data['empleados_ids']) == 0,
            'procesados': data['procesados'],
            'total': data['total'],
            'url_retorno': f"/admin/{data['app_label']}/{data['model_name']}/"
        })

    # Renderizado inicial de la Pantalla de Carga limpia (HTML Puro)
    html_content = f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Enviando Evaluaciones...</title>
        <style>
            body {{
                margin: 0; padding: 0; width: 100vw; height: 100vh;
                background-color: #f3f4f6;
                display: flex; justify-content: center; align-items: center;
                font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
            }}
            .card {{
                background: #ffffff; padding: 35px 50px; border-radius: 12px;
                box-shadow: 0 20px 25px -5px rgba(0,0,0,0.1); text-align: center;
                border: 1px solid #e5e7eb; max-width: 420px;
            }}
            .spinner {{
                animation: spin 1s linear infinite; margin: 0 auto 15px auto;
                width: 42px; height: 42px; color: #72a651;
            }}
            .progress-bar-container {{
                width: 100%; background-color: #e5e7eb; border-radius: 9999px; height: 8px; margin-top: 15px; overflow: hidden;
            }}
            .progress-bar {{
                width: 0%; height: 100%; background-color: #72a651; transition: width 0.3s ease;
            }}
            @keyframes spin {{ to {{ transform: rotate(360deg); }} }}
        </style>
    </head>
    <body>
        <div class="card">
            <svg class="spinner" fill="none" viewBox="0 0 24 24" xmlns="http://www.w3.org/2000/svg">
                <circle cx="12" cy="12" r="10" stroke="currentColor" stroke-width="4" style="opacity: 0.25;"></circle>
                <path fill="currentColor" d="M4 12a8 8 0 018-8V0C5.373 0 0 5.373 0 12h4zm2 5.291A7.962 7.962 0 014 12H0c0 3.042 1.135 5.824 3 7.938l3-2.647z" style="opacity: 0.75;"></path>
            </svg>
            <h3 style="margin: 0 0 8px 0; font-size: 18px; font-weight: 600; color: #1f2937;">
                Enviando Enlaces de Evaluación
            </h3>
            <p id="status-text" style="margin: 0; font-size: 14px; color: #6b7280; line-height: 1.4;">
                Iniciando el envío seguro de correos...
            </p>
            <div class="progress-bar-container">
                <div id="progress" class="progress-bar"></div>
            </div>
        </div>

        <script>
            function realizarEnvio() {{
                fetch(window.location.pathname + "?ajax=1", {{
                    headers: {{ 'X-Requested-With': 'XMLHttpRequest' }}
                }})
                .then(res => res.json())
                .then(data => {{
                    let porcentaje = Math.round((data.procesados / data.total) * 100);
                    document.getElementById("progress").style.width = porcentaje + "%";
                    document.getElementById("status-text").innerText = "Enviados: " + data.procesados + " de " + data.total + " correos electrónicos...";
                    
                    if (data.completado) {{
                        window.location.href = data.url_retorno;
                    }} else {{
                        // Pequeño delay de 100ms para no saturar el servidor y continuar el bucle
                        setTimeout(realizarEnvio, 100);
                    }}
                }})
                .catch(() => {{
                    alert("Ocurrió un problema en el servidor al enviar los correos.");
                }});
            }}
            // Iniciar ciclo en cuanto carge la vista
            window.onload = realizarEnvio;
        </script>
    </body>
    </html>
    """
    return HttpResponse(html_content)

#admin.site.register(Departamento, DepartamentoAdmin)
admin_site.register(Departamento, DepartamentoAdmin)
class EmpleadoAdmin(CatalogosOrdenadosAdmin, ExcelImportAdmin):
    form = EmpleadoAdminForm  
    model_class = Empleado
    pk_field_name = 'id_empleado'
    excel_columns = ['id_empleado', 'nombre_largo', 'id_puesto_id', 'id_departamento_id', 'CorreoElectronico', 'estado_empleado', 'fechaalta', 'se_evalua']
    list_display = ('id_empleado', 'nombre_largo', 'id_puesto', 'id_departamento', 'id_jefe', 'es_jefe_departamento', 'CorreoElectronico', 'estado_empleado', 'fechaalta', 'se_evalua', 'acciones_rh')
    list_filter = ('id_departamento', 'id_puesto', 'es_jefe_departamento', 'CorreoElectronico', 'id_jefe', 'estado_empleado', 'se_evalua', 'id_jefe__nombre_largo')
    search_fields = ('nombre_largo', 'id_puesto__descripcion', 'CorreoElectronico', 'id_jefe__nombre_largo')

# 🌟 ÚNICA DEFINICIÓN:
    action_form = CustomActionForm
    action_submit_label = "Ejecutar"

    inlines = []  
    actions = [enviar_enlaces_seleccionados]
    # 🌟 AGREGA ESTE MÉTODO JUSTO AQUÍ ADENTRO:

    def change_view(self, request, object_id, form_url='', extra_context=None):
        extra_context = extra_context or {}
        obj = self.get_object(request, object_id)
        matriz_html = ""
        
        if obj:
            # 1. Instanciamos tu formulario personalizado
            form = EmpleadoAdminForm(instance=obj)
            
            # 🌟 ACCIÓN CLAVE: Forzamos al campo interno a aceptar TODAS las competencias unificadas
            queryset = form.queryset_competencias_custom
            form.fields['competencias_seleccionadas'].queryset = queryset
            
            # 2. Obtenemos el set de IDs guardados utilizando la columna relacional física
            guardados_ids = set(EmpleadoCompetenciaAsignada.objects.filter(
                id_empleado_id=obj.pk
            ).values_list('id_competencia_id', flat=True))

            if queryset.exists():
                matriz_html = '''
                <style>
                    .matriz-unfold-container {
                        width: 100% !important;
                        margin-top: 2rem !important;
                        clear: both !important;
                    }
                </style>
                '''
                
                matriz_html += '<div class="matriz-unfold-container bg-gray-50 dark:bg-zinc-900/40 p-6 rounded-lg border border-gray-200 dark:border-zinc-800">'
                matriz_html += '<h2 class="text-base font-semibold text-gray-900 dark:text-white mb-1">Asignación de Competencias Específicas</h2>'
                matriz_html += '<p class="text-sm text-gray-500 dark:text-zinc-400 mb-6">Palomee las competencias específicas del puesto o libres que formarán parte de la evaluación individual de este colaborador.</p>'
                
                ultima_clasificacion = None

                # 3. Iteramos exactamente sobre el listado del UNION
                for comp in queryset:
                    clasif_nombre = comp.id_clasificacion.descripcion if comp.id_clasificacion else "Competencias Específicas Sueltas"
                    
                    if clasif_nombre != ultima_clasificacion:
                        if ultima_clasificacion is not None:
                            matriz_html += '</div></div>' 
                        
                        matriz_html += f'''
                        <div class="mb-6">
                            <h3 class="text-xs font-semibold uppercase tracking-wider text-blue-600 dark:text-blue-400 mb-3 border-b border-gray-200 dark:border-zinc-800 pb-2 flex items-center gap-2">
                                <span>📂</span> {clasif_nombre}
                            </h3>
                            <div class="grid grid-cols-1 gap-2">
                        '''
                        ultima_clasificacion = clasif_nombre

                    # 4. Evaluamos si el ID numérico de la competencia iterada está guardado
                    is_checked = comp.id_competencia in guardados_ids
                    checked_str = "checked" if is_checked else ""

                    matriz_html += f'''
                        <label class="flex items-center gap-4 bg-white dark:bg-zinc-800/40 hover:bg-gray-100 dark:hover:bg-zinc-800 px-4 py-3 rounded-md border border-gray-200 dark:border-zinc-800/80 cursor-pointer transition-all w-full block">
                            <input type="checkbox" name="competencias_seleccionadas" value="{comp.id_competencia}" {checked_str} class="rounded border-gray-300 dark:border-zinc-700 text-blue-600 focus:ring-blue-500 h-4 w-4" style="accent-color: #3b82f6; min-width: 16px;">
                            <span class="text-gray-700 dark:text-zinc-300 text-sm font-normal leading-normal">{comp.descripcion}</span>
                        </label>
                    '''

                matriz_html += '</div></div></div>'
            else:
                matriz_html = '<div class="mt-6 bg-gray-50 dark:bg-zinc-900/40 p-4 rounded-lg border border-gray-200 dark:border-zinc-800"><p class="text-sm italic text-gray-400">No hay competencias específicas configuradas para el puesto de este empleado.</p></div>'

        extra_context['matriz_competencias_html'] = mark_safe(matriz_html)
        
        # 🌟 PASO COMPLEMENTARIO OBLIGATORIO PARA UNFOLD:
        # Inyectamos el objeto de formulario procesado en el contexto para que reemplace al predeterminado
        extra_context['adminform'] = admin.helpers.AdminForm(
            form,
            list(self.get_fieldsets(request, obj)),
            self.get_prepopulated_fields(request, obj),
            self.get_readonly_fields(request, obj),
            model_admin=self
        )
        
        return super().change_view(request, object_id, form_url, extra_context=extra_context)

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        
        competencias_post = request.POST.getlist('competencias_seleccionadas')
        competencias_post_ids = [int(pk) for pk in competencias_post if pk.isdigit()]

        # 1. Eliminar desmarcados
        EmpleadoCompetenciaAsignada.objects.filter(id_empleado=obj).exclude(id_competencia_id__in=competencias_post_ids).delete()

        # 2. Guardar nuevos checks marcados
        for comp_id in competencias_post_ids:
            EmpleadoCompetenciaAsignada.objects.get_or_create(id_empleado=obj, id_competencia_id=comp_id)

#admin.site.register(Empleado, EmpleadoAdmin)
admin_site.register(Empleado, EmpleadoAdmin)

class CompetenciaClasificacionAdmin(ExcelImportAdmin):
    model_class = CompetenciaClasificacion
    pk_field_name = 'id_clasificacion'
    excel_columns = ['descripcion', 'tipo']
    list_display = ('id_clasificacion', 'descripcion', 'tipo', 'acciones_rh')
    list_filter = ('descripcion', 'tipo')
    search_fields = ('descripcion',)
    fields = ('descripcion', 'tipo')
    inlines = [CompetenciaInline]

#admin.site.register(CompetenciaClasificacion, CompetenciaClasificacionAdmin)
admin_site.register(CompetenciaClasificacion, CompetenciaClasificacionAdmin)

class CompetenciaAdmin(CatalogosOrdenadosAdmin, ExcelImportAdmin):
    model_class = Competencia
    pk_field_name = 'id_competencia'
    excel_columns = ['id_clasificacion_id', 'descripcion']
    list_display = ('id_competencia', 'id_clasificacion', 'descripcion', 'acciones_rh')
    list_filter = ('descripcion',)  
    search_fields = ('descripcion',)

#admin.site.register(Competencia, CompetenciaAdmin)
admin_site.register(Competencia, CompetenciaAdmin)


class EvaluacionAdmin(ExcelImportAdmin):
    model_class = Evaluacion
    pk_field_name = 'id_evaluacion'
    excel_columns = ['descripcion', 'fecha_inicial', 'fecha_final']
    list_display = ('id_evaluacion', 'descripcion', 'fecha_inicial', 'fecha_final', 'acciones_rh', 'cerrada')
    search_fields = ('descripcion',)

    # 🌟 ÚNICA DEFINICIÓN:
    action_form = CustomActionForm
    action_submit_label = "Ejecutar"    
    actions = [enviar_enlaces_magicos]

#admin.site.register(Evaluacion, EvaluacionAdmin)
admin_site.register(Evaluacion, EvaluacionAdmin)